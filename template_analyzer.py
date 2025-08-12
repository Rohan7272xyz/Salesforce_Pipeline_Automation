import os
import re
from pathlib import Path
from openpyxl import load_workbook
import sys
from datetime import datetime

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

class TemplateAnalyzer:
    """Analyzes Excel template and generates a complete new app.py from template"""
    
    def __init__(self):
        self.template_path = Config.TEMPLATE_PATH
        self.app_py_path = project_root / "app.py"
        self.backup_dir = Config.BACKUP_DIR
        
    def extract_template_headers(self):
        """Extract column headers from the template file."""
        try:
            print(f"📋 Analyzing template: {self.template_path}")
            
            workbook = load_workbook(self.template_path)
            sheet = workbook[Config.TEMPLATE_SHEET_NAME]
            
            # Headers are typically in row 4 (Config.DATA_START_ROW - 1)
            header_row = Config.DATA_START_ROW - 1
            headers = []
            
            # Extract headers and track their actual Excel column positions
            empty_count = 0
            for col in range(1, 50):  # Check first 50 columns max
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value is None or str(cell_value).strip() == "":
                    empty_count += 1
                    # Stop if we hit 3 consecutive empty cells (allows for blank Column A)
                    if empty_count >= 3:
                        break
                    continue
                else:
                    empty_count = 0  # Reset counter when we find content
                    headers.append({
                        'name': str(cell_value).strip(),
                        'excel_position': col  # Track actual Excel column position
                    })
            
            print(f"✅ Found {len(headers)} total columns in template:")
            for i, header in enumerate(headers):
                print(f"   Excel Col {header['excel_position']}: {header['name']}")
                
            return headers
            
        except Exception as e:
            print(f"❌ Error analyzing template: {e}")
            raise
    
    def identify_input_vs_calendar_columns(self, headers):
        """
        Separate input data columns (from Salesforce) from calendar/forecast columns.
        
        Input columns: All Salesforce fields (everything except calendar dates)
        Calendar columns: Only monthly forecast columns (Jun 2025, Jul 2025, etc.)
        """
        input_columns = []
        calendar_columns = []
        
        # Define strict patterns for calendar/forecast columns - ONLY date patterns
        month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                      'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        
        for header_info in headers:
            header_name = header_info['name']
            excel_position = header_info['excel_position']
            header_lower = header_name.lower().strip()
            
            # Only treat as calendar if it matches month + year pattern
            is_calendar = False
            
            # Check for month + year patterns (e.g., "Jun 2025", "jul 2026")
            for month in month_names:
                # Pattern: "Month Year" (e.g., "Jun 2025", "Jul 2026") 
                if (month in header_lower and 
                    any(year in header_lower for year in ['2025', '2026', '2027', '2028', '2029', '2030'])):
                    # Additional check: make sure it's actually a month-year format
                    # and not just coincidental text containing month name + year
                    words = header_lower.split()
                    if (len(words) == 2 and 
                        words[0] in month_names and 
                        words[1].isdigit() and 
                        int(words[1]) >= 2025):
                        is_calendar = True
                        break
            
            if is_calendar:
                calendar_columns.append({
                    'index': excel_position,
                    'name': header_name,
                    'type': 'calendar'
                })
            else:
                # Everything else is input data (including T&E, Contract fields, etc.)
                input_columns.append({
                    'index': excel_position,
                    'name': header_name,
                    'type': self.determine_column_type(header_name)
                })
        
        print(f"📊 Column Analysis:")
        print(f"   Input/Data columns: {len(input_columns)}")
        print(f"   Calendar/Forecast columns: {len(calendar_columns)}")
        
        print(f"\n📋 Input columns identified:")
        for info in input_columns:
            print(f"   - {info['name']} (Excel position {info['index']}, type: {info['type']})")
            
        print(f"\n📅 Calendar columns identified:")
        for info in calendar_columns:
            print(f"   - {info['name']} (Excel position {info['index']})")
        
        return input_columns, calendar_columns
    
    def determine_column_type(self, header):
        """Analyze header to determine data type and formatting."""
        header_lower = header.lower()
    
        # Specific fixes for the three problematic columns
        if 'positioning' in header_lower:
            return 'text'  # Force Positioning to be treated as text
        elif 'govwin' in header_lower:
            return 'text'  # Force GovWin to be treated as text
        elif 'rfp' in header_lower and 'award' in header_lower:
            return 'date'  # Force RFP Award to be treated as date
        # Existing logic continues below
        elif any(keyword in header_lower for keyword in ['value', 'ceiling', 'contract', 'price', 'cost', 'amount']) and '$' in header:
            return 'currency'
        elif any(keyword in header_lower for keyword in ['date', 'deadline', 'due']):
            return 'date'
        elif any(keyword in header_lower for keyword in ['win', 'probability', 'percent', '%']):
            return 'percentage'
        elif any(keyword in header_lower for keyword in ['opportunity', 'description', 'notes', 'name']):
            return 'text_wrap'
        else:
            return 'text'

    def analyze_raw_data_sample(self):
        """
        Analyze the most recent raw data file to understand column structure.
        """
        try:
            # Look for recent input files to analyze
            input_files = list(Config.INPUT_DIR.glob("pipeline_*.xlsx"))
            if not input_files:
                print("⚠️ No sample input files found for smart mapping analysis")
                return None
            
            # Use the most recent file
            latest_file = max(input_files, key=lambda f: f.stat().st_mtime)
            print(f"🔍 Analyzing raw data structure from: {latest_file.name}")
            
            workbook = load_workbook(latest_file, data_only=True)
            sheet = workbook.sheetnames[0]
            raw_sheet = workbook[sheet]
            
            # Get first few rows to find headers
            raw_values = []
            for row in raw_sheet.iter_rows(values_only=True, min_row=1, max_row=20):
                raw_values.append([str(cell).strip() if cell is not None else "" for cell in row])
            
            # Find header row (typically row 14, 0-indexed = 13)
            header_row_data = None
            for start_row in range(10, len(raw_values)):
                row = raw_values[start_row]
                row_text = " ".join(str(cell).lower() for cell in row if cell and cell.strip())
                
                # Look for key Salesforce column indicators
                indicators = ['capture', 'opportunity', 'salesforce', 'stage', 'positioning', 'govwin']
                matches = sum(1 for indicator in indicators if indicator in row_text)
                
                if matches >= 3:
                    print(f"📋 Found raw data headers at row {start_row + 1}")
                    header_row_data = row
                    data_start_row = start_row + 1  # Data starts one row after headers
                    break
            
            if not header_row_data:
                print("⚠️ Could not identify header row, using row 14 as fallback")
                header_row_data = raw_values[13] if len(raw_values) > 13 else None
                data_start_row = 14
            
            if not header_row_data:
                return None
            
            # Get sample data rows
            sample_data_rows = []
            for i in range(data_start_row, min(data_start_row + 5, len(raw_values))):
                if i < len(raw_values):
                    sample_data_rows.append(raw_values[i])
            
            print(f"📊 Raw data structure analysis:")
            print(f"   Headers: {[h for h in header_row_data if h and h.strip()]}")
            print(f"   Sample data rows: {len(sample_data_rows)}")
            
            return {
                'headers': header_row_data,
                'sample_data': sample_data_rows,
                'data_start_row': data_start_row
            }
            
        except Exception as e:
            print(f"⚠️ Error analyzing raw data sample: {e}")
            import traceback
            print(traceback.format_exc())
            return None
    
    def smart_column_mapping(self, template_columns, raw_data_info):
        """
        Create smart mapping between template columns and raw data columns.
        Uses actual data content analysis, not just header names.
        """
        if not raw_data_info:
            return None
        
        raw_headers = raw_data_info['headers']
        sample_data = raw_data_info['sample_data']
        
        print(f"🔗 Smart Column Mapping Analysis:")
        print(f"   Template columns: {len(template_columns)}")
        print(f"   Raw data columns: {len(raw_headers)}")
        
        # Create mapping from template column to raw column index
        template_to_raw_mapping = {}
        
        for template_col in template_columns:
            template_name = template_col['name']
            template_lower = template_name.lower().strip()
            
            best_match_index = None
            best_match_score = 0
            
            # Analyze each raw column
            for raw_idx, raw_header in enumerate(raw_headers):
                if not raw_header or not raw_header.strip():
                    continue
                    
                raw_header_lower = raw_header.lower().strip()
                
                # Score this column based on multiple factors
                score = 0
                
                # 1. Header name similarity
                if template_lower in raw_header_lower or raw_header_lower in template_lower:
                    score += 10
                
                # 2. Specific keyword matching
                if 'positioning' in template_lower and 'positioning' in raw_header_lower:
                    score += 20
                elif 'govwin' in template_lower and 'govwin' in raw_header_lower:
                    score += 20
                elif 'capture' in template_lower and 'capture' in raw_header_lower:
                    score += 15
                elif 'opportunity' in template_lower and 'opportunity' in raw_header_lower:
                    score += 15
                elif ('sf' in template_lower or 'salesforce' in template_lower) and ('salesforce' in raw_header_lower or 'sf' in raw_header_lower):
                    score += 20
                elif 'award' in template_lower and 'award' in raw_header_lower:
                    score += 15
                elif 'ceiling' in template_lower and 'ceiling' in raw_header_lower:
                    score += 15
                elif 'stage' in template_lower and 'stage' in raw_header_lower:
                    score += 15
                
                # 3. Data content analysis
                if sample_data:
                    sample_values = [row[raw_idx] if raw_idx < len(row) else '' for row in sample_data[:3]]
                    sample_values = [str(v).strip() for v in sample_values if v]
                    
                    if sample_values:
                        # Check data patterns to confirm mapping
                        if 'positioning' in template_lower:
                            # Should contain stage-like text (Sub, Capture, Qualification)
                            if any('sub' in v.lower() or 'capture' in v.lower() or 'qualification' in v.lower() for v in sample_values):
                                score += 30
                        elif 'govwin' in template_lower:
                            # Should contain numeric IDs
                            if any(v.isdigit() and len(v) >= 5 for v in sample_values):
                                score += 30
                        elif 'award' in template_lower or 'rfp' in template_lower:
                            # Should contain dates
                            if any('/' in v and any(c.isdigit() for c in v) for v in sample_values):
                                score += 25
                        elif 'ceiling' in template_lower or 'value' in template_lower:
                            # Should contain large numbers
                            if any(v.replace(',', '').replace('$', '').isdigit() for v in sample_values):
                                score += 20
                
                if score > best_match_score:
                    best_match_score = score
                    best_match_index = raw_idx
            
            if best_match_index is not None:
                template_to_raw_mapping[template_name] = {
                    'raw_index': best_match_index,
                    'raw_header': raw_headers[best_match_index],
                    'score': best_match_score
                }
                print(f"   ✅ {template_name} → Raw Column {best_match_index + 1} '{raw_headers[best_match_index]}' (score: {best_match_score})")
            else:
                print(f"   ❌ No mapping found for: {template_name}")
        
        return template_to_raw_mapping
    
    def generate_app_py_content(self, input_columns, calendar_columns, column_mapping):
        """Generate complete app.py content from template structure with smart column mapping."""
        
        if not column_mapping:
            print("❌ No column mapping available, cannot generate app.py")
            return None
        
        # Only use input columns for processing - calendar columns are template-only
        processing_columns = input_columns
        all_template_columns = input_columns + calendar_columns
        
        # Generate expected_columns list based on actual raw data positions
        # After dropping column 0, raw column positions become: position - 1
        expected_columns_list = []
        mapped_column_names = []
        
        for col_info in processing_columns:
            template_name = col_info['name']
            if template_name in column_mapping:
                raw_index = column_mapping[template_name]['raw_index']
                # After dropping column 0: raw_index stays the same (since we drop index 0)
                expected_columns_list.append(str(raw_index))
                mapped_column_names.append(f"'{template_name}'")  # Use template names
            else:
                print(f"⚠️ Warning: No mapping found for {template_name}, skipping")
                continue
        
        if len(expected_columns_list) == 0:
            print("❌ No valid column mappings found!")
            return None
        expected_columns_str = ', '.join(expected_columns_list)
        column_names_list = ',\n    '.join(mapped_column_names)
        
        # Generate data processing code (only for input columns)
        data_processing_lines = []
        for info in processing_columns:
            if info['type'] == 'currency':
                data_processing_lines.append(f"        df_raw['{info['name']}'] = pd.to_numeric(df_raw['{info['name']}'].str.replace(r'[\\$,]', '', regex=True), errors='coerce')")
            elif info['type'] == 'percentage':
                data_processing_lines.append(f"        df_raw['{info['name']}'] = pd.to_numeric(df_raw['{info['name']}'], errors='coerce')")
            elif info['type'] == 'number':
                data_processing_lines.append(f"        df_raw['{info['name']}'] = pd.to_numeric(df_raw['{info['name']}'], errors='coerce')")

        data_processing_code = '\n'.join(data_processing_lines) if data_processing_lines else '        # No special data processing needed'

        # Generate cell formatting code (for all template columns, using CORRECTED template positions)
        cell_formatting_lines = []
        text_wrap_checks = []
        
        for info in all_template_columns:
            # IMPORTANT: Add +1 to Excel position because raw data Column A is always blank
            # This ensures data goes to the correct template columns
            j = info['index'] + 1  # Shift Excel position by +1
            name = info['name']  # Use template name for comments
            col_type = info['type']
            
            if col_type == 'currency':
                cell_formatting_lines.append(f"""                elif j == {j}:  # {name}
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val""")
            elif col_type == 'date':
                cell_formatting_lines.append(f"""                elif j == {j}:  # {name}
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'""")
                
            elif col_type == 'number':
                cell_formatting_lines.append(f"""                elif j == {j}:  # {name}
                    try:
                        cell.value = int(float(val))
                        cell.number_format = '0'
                    except (ValueError, TypeError):
                        cell.value = val""")

            elif col_type == 'percentage':
                cell_formatting_lines.append(f"""                elif j == {j}:  # {name}
                    try:
                        cell.value = int(float(val))
                        cell.number_format = '0'
                    except (ValueError, TypeError):
                        cell.value = val""")
            
            if col_type == 'text_wrap':
                text_wrap_checks.append(f"                if j == {j}:  # {name}")
                text_wrap_checks.append(f"                    cell.alignment = Alignment(wrap_text=True, vertical='top')")

        cell_formatting_code = '\n'.join(cell_formatting_lines) if cell_formatting_lines else '                # No special formatting needed'
        text_wrap_code = '\n'.join(text_wrap_checks) if text_wrap_checks else '                # No text wrapping needed'
        
        # Total number of template columns for clearing
        total_template_columns = len(all_template_columns)
        
        # Input columns count for processing
        input_columns_count = len(processing_columns)
        
        # Create mapping info for documentation
        mapping_info = "\\n".join([
            f"   {info['name']} → Raw Column {column_mapping.get(info['name'], {}).get('raw_index', 'UNKNOWN')} '{column_mapping.get(info['name'], {}).get('raw_header', 'UNKNOWN')}'"
            if isinstance(column_mapping.get(info['name'], {}).get('raw_index'), int)
            else f"   {info['name']} → NO MAPPING FOUND"
            for info in processing_columns
        ])
        
        # Generate the complete app.py content
        app_py_content = f'''import pandas as pd
import os
import eel
import tkinter as tk
from openpyxl.styles import Alignment
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

# --- EEL Setup ---
eel.init(str(Config.WEB_DIR))

# --- Logging Shim ---
def safe_log(message, level='info'):
    """Logs messages to the eel frontend if available, otherwise prints to console."""
    if hasattr(eel, 'logMessage'):
        try:
            eel.logMessage(message, level)
            return
        except Exception:
            pass
    prefix = level.upper()
    print(f"[{{prefix}}] {{message}}")

def safe_complete(ok: bool):
    """Signals completion status to the eel frontend if available."""
    if hasattr(eel, 'processingComplete'):
        try:
            eel.processingComplete(ok)
            return
        except Exception:
            pass
    print(f"[STATUS] processingComplete({{ok}})")

@eel.expose
def select_file(title, file_type):
    """Opens a native file dialog to select a file."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_path = filedialog.askopenfilename(
            title=title, filetypes=[(file_type, "*.xlsx *.xls")]
        )
        return file_path
    except Exception as e:
        safe_log(f"Error in file selection: {{e}}", 'error')
        return ""

# --- Date parsing helper ---
def parse_date(val):
    """Safely parses a value to a date object, returning original value on failure."""
    try:
        return pd.to_datetime(val).date()
    except:
        return val

@eel.expose
def process_and_merge_files(data_path):
    """
    Main function to process the raw data file, merge it into a template,
    and save the output. It ensures the 'Total' row is always last.
    Auto-generated with smart column mapping.
    
    Input Data Columns: {input_columns_count}
    Total Template Columns: {total_template_columns}
    
    Smart Column Mapping:
{mapping_info}
    """
    try:
        safe_log("--- Starting Excel Processing ---")

        if not data_path or not Path(data_path).exists():
            safe_log("Error: Raw Data File not found or not provided.", 'error')
            safe_complete(False)
            return None

        if not Config.TEMPLATE_PATH.exists():
            safe_log(f"Error: Integrated template file not found at {{Config.TEMPLATE_PATH}}", 'error')
            safe_complete(False)
            return None

        # --- Strip Formatting and Metadata ---
        safe_log(f"Reading and cleaning data from '{{Path(data_path).name}}'...")
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        raw_values = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in sheet_raw.iter_rows(values_only=True)
        ]

        df_raw = pd.DataFrame(raw_values)
        df_raw = df_raw.iloc[14:].copy()  # Data starts at row 14 (0-indexed)
        df_raw.drop(columns=[0], inplace=True, errors='ignore')  # Remove empty first column
        df_raw.reset_index(drop=True, inplace=True)

        # --- Input Data Column Structure (Salesforce fields only) ---
        expected_columns = [{expected_columns_str}]
        if any(idx >= len(df_raw.columns) for idx in expected_columns):
            safe_log(f"Error: Expected {{len(expected_columns)}} columns but input has {{len(df_raw.columns)}} columns.", 'error')
            safe_log("Input columns needed: " + str(expected_columns), 'error')
            safe_log("Available columns: " + str(list(range(len(df_raw.columns)))), 'error')
            safe_complete(False)
            return None

        df_raw = df_raw[expected_columns]
        df_raw.columns = [
    {column_names_list}
]

        # DEBUG: Check if the problematic columns have data
        safe_log("DEBUG: Column names after mapping:")
        safe_log(str(df_raw.columns.tolist()))
        safe_log("DEBUG: Sample data from first row:")
        for col in df_raw.columns:
            sample_val = df_raw[col].iloc[0] if len(df_raw) > 0 else "NO DATA"
            safe_log(f"  {{col}}: '{{sample_val}}'")
        safe_log("DEBUG: Check specific columns:")
        for col_name in ['Award Date', 'GovWin IQ Opportunity ID', 'Positioning']:
            if col_name in df_raw.columns:
                non_empty = df_raw[col_name].dropna()
                non_empty = non_empty[non_empty != '']
                safe_log(f"  {{col_name}}: {{len(non_empty)}} non-empty values out of {{len(df_raw)}}")
                if len(non_empty) > 0:
                    safe_log(f"    Sample values: {{non_empty.head(3).tolist()}}")
            else:
                safe_log(f"  {{col_name}}: COLUMN NOT FOUND!")

        # --- Auto-generated Data Processing (Input columns only) ---
{data_processing_code}

        df = df_raw.dropna(how='all')
        df.dropna(subset=[df.columns[1]], inplace=True)  # Use second column for opportunity check

        # --- Exclude unwanted text across multiple columns ---
        EXCLUSION_KEYWORDS = [
            'Confidential Information - Do Not Distribute',
            'Copyright © 2000-2025 salesforce.com, inc. All rights reserved.'
        ]
        for col in [df.columns[0], df.columns[1]]:  # Check first two columns
            df = df[~df[col].astype(str).str.contains('|'.join(EXCLUSION_KEYWORDS), case=False, na=False)]

        # --- Handle Total Row AFTER filtering ---
        safe_log("Separating and sorting main data from total row...")

        # Create a temporary column for case-insensitive matching and trimming
        df['temp_mgr_lower'] = df[df.columns[0]].astype(str).str.strip().str.lower()

        # Isolate the total row(s) and the main data
        total_row_mask = df['temp_mgr_lower'] == 'total'
        total_df = df[total_row_mask].copy()
        main_df = df[~total_row_mask].copy()

        # Handle stray total count rows
        if not total_df.empty:
            count_row_mask = main_df[df.columns[1]].astype(str).str.match(r'^\\d+$')
            if count_row_mask.any():
                numeric_value = main_df.loc[count_row_mask, df.columns[1]].iloc[0]
                total_df.iloc[0, total_df.columns.get_loc(df.columns[1])] = numeric_value
                main_df = main_df[~count_row_mask]
                safe_log("Found and moved stray total count to the total row.")

        # --- Sort data in the desired order: by manager, then unassigned, then total ---
        safe_log("Sorting data: assigned managers, unassigned, then total...")

        has_mgr_mask = (main_df['temp_mgr_lower'] != '') & (main_df['temp_mgr_lower'] != 'nan')
        df_with_mgr = main_df[has_mgr_mask].copy()
        df_without_mgr = main_df[~has_mgr_mask].copy()

        df_with_mgr.sort_values(by=df.columns[0], inplace=True)

        df = pd.concat([df_with_mgr, df_without_mgr, total_df], ignore_index=True)
        df.drop(columns=['temp_mgr_lower'], inplace=True, errors='ignore')

        safe_log(f"Processed {{len(df)}} rows of data.")

        # DEBUG: Check DataFrame just before writing to Excel
        safe_log("DEBUG: Final DataFrame columns before Excel writing:")
        safe_log(f"  DataFrame shape: {{df.shape}}")
        safe_log(f"  Columns with data for problem columns:")
        for col_name in ['Award Date', 'GovWin IQ Opportunity ID', 'Positioning']:
            if col_name in df.columns:
                col_index = df.columns.get_loc(col_name)
                non_empty = df[col_name].dropna()
                non_empty = non_empty[non_empty != '']
                safe_log(f"    {{col_name}} (pandas index {{col_index}}): {{len(non_empty)}} values")
            else:
                safe_log(f"    {{col_name}}: NOT FOUND IN DATAFRAME")

        # --- Load Template ---
        safe_log("Loading integrated template file...")
        workbook = load_workbook(Config.TEMPLATE_PATH)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]

        # --- Clear Old Data (All template columns) ---
        end_row = sheet.max_row
        for r in range(Config.DATA_START_ROW, end_row + 1):
            val = sheet.cell(row=r, column=1).value
            if isinstance(val, str) and "total" in val.lower():
                end_row = r - 1
                break

        # Clear all template columns (input + calendar) - add +1 for correct positioning
        for r_idx in range(Config.DATA_START_ROW, end_row + 1):
            for c_idx in range(2, {total_template_columns + 2}):  # Start from column 2 (B), not 1 (A)
                sheet.cell(row=r_idx, column=c_idx).value = None

        safe_log(f"Cleared rows {{Config.DATA_START_ROW}} to {{end_row}} in the template.")

        # --- Write New Data (Only input columns have data) ---
        safe_log("Writing new data to template...")
        for i, row in enumerate(df.itertuples(index=False), start=Config.DATA_START_ROW):
            sheet.row_dimensions[i].height = 30  # Set uniform row height

            for j, val in enumerate(row, start=2):  # START FROM COLUMN 2 (B), NOT 1 (A)
                cell = sheet.cell(row=i, column=j)

                if pd.isna(val):
                    cell.value = None
{cell_formatting_code}
                else:
                    cell.value = val

                # Text wrapping for specific columns
{text_wrap_code}

        downloads_path = Config.get_downloads_path()
        output_filename = Config.DEFAULT_OUTPUT_NAME
        output_path = downloads_path / output_filename

        workbook.save(str(output_path))
        safe_log(f"Success! Final file saved to '{{output_path}}'.", 'success')
        safe_complete(True)
        return str(output_path)

    except Exception as e:
        safe_log(f"An unexpected error occurred: {{e}}", 'error')
        import traceback
        safe_log(traceback.format_exc(), 'error')
        safe_complete(False)
        return None

if __name__ == "__main__":
    # Validate configuration first
    try:
        Config.validate_config()
        print("Configuration validated successfully")
    except ValueError as e:
        print(f"Configuration error: {{e}}")
        sys.exit(1)
    
    # This allows running the script directly for testing or from the command line
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
        result_path = process_and_merge_files(data_path)
        if result_path:
            print(f"[RESULT] {{result_path}}")
        else:
            sys.exit(1)
    else:
        # This starts the eel GUI application
        try:
            eel.start('index.html', port=0, cmdline_args=['--start-maximized'])
        except (SystemExit, MemoryError, KeyboardInterrupt):
            print("Application closed.")
        except Exception as e:
            print(f"Failed to start GUI: {{e}}")
            print("Make sure the 'web' directory exists with index.html")
            sys.exit(1)
'''
        
        return app_py_content
    
    def backup_current_app_py(self):
        """Create a backup of the current app.py before replacing it."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"app_py_backup_{timestamp}.py"
            backup_path = self.backup_dir / backup_filename
            
            # Ensure backup directory exists
            self.backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Copy current app.py to backup
            with open(self.app_py_path, 'r', encoding='utf-8') as source:
                with open(backup_path, 'w', encoding='utf-8') as backup:
                    backup.write(source.read())
            
            print(f"📁 Backed up current app.py to: {backup_path}")
            return backup_path
            
        except Exception as e:
            print(f"⚠️ Warning: Could not backup app.py: {e}")
            return None
    
    def write_new_app_py(self, new_content):
        """Write completely new app.py content."""
        try:
            with open(self.app_py_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            print(f"✅ Successfully generated new app.py")
            return True
            
        except Exception as e:
            print(f"❌ Error writing new app.py: {e}")
            return False
    
    def analyze_and_update(self):
        """Main function to analyze template and generate new app.py."""
        try:
            print("🔍 Starting smart template analysis and app.py regeneration...")
            
            # Step 1: Extract headers from template
            headers = self.extract_template_headers()
            if not headers:
                raise ValueError("No headers found in template")
            
            # Step 2: Separate input columns from calendar columns
            input_columns, calendar_columns = self.identify_input_vs_calendar_columns(headers)
            
            if not input_columns:
                raise ValueError("No input data columns identified")
            
            # Step 3: Analyze raw data structure
            raw_data_info = self.analyze_raw_data_sample()
            if not raw_data_info:
                print("⚠️ Warning: Could not analyze raw data, using template column order")
                column_mapping = None
            else:
                # Step 4: Create smart column mapping
                column_mapping = self.smart_column_mapping(input_columns, raw_data_info)
            
            # Step 5: Generate complete new app.py content
            new_app_content = self.generate_app_py_content(input_columns, calendar_columns, column_mapping)
            
            if not new_app_content:
                raise ValueError("Failed to generate app.py content")
            
            # Step 6: Backup current app.py
            self.backup_current_app_py()
            
            # Step 7: Write new app.py
            success = self.write_new_app_py(new_app_content)
            
            if success:
                print("🎉 Smart template analysis complete! New app.py generated.")
                print(f"📊 Processing Structure:")
                print(f"   Input data columns: {len(input_columns)} (from Salesforce)")
                print(f"   Calendar columns: {len(calendar_columns)} (template only)")
                print(f"   Total template columns: {len(input_columns + calendar_columns)}")
                
                if column_mapping:
                    print("\n📋 Column mappings applied:")
                    for template_col in input_columns:
                        if template_col['name'] in column_mapping:
                            mapping = column_mapping[template_col['name']]
                            print(f"   - {template_col['name']} ← Raw Column {mapping['raw_index'] + 1} '{mapping['raw_header']}'")
                
                return True
            else:
                print("❌ Failed to write new app.py")
                return False
                
        except Exception as e:
            print(f"❌ Template analysis failed: {e}")
            import traceback
            print(traceback.format_exc())
            return False

def analyze_template_and_update_app():
    """Main entry point for template analysis - called by other scripts."""
    analyzer = TemplateAnalyzer()
    return analyzer.analyze_and_update()

if __name__ == "__main__":
    try:
        Config.validate_config()
        print("Configuration validated")
    except ValueError as e:
        print(f"Configuration error: {e}")
        sys.exit(1)
    
    # Run the analysis
    success = analyze_template_and_update_app()
    
    if success:
        print("\n✅ Smart template analysis completed successfully!")
        print("🚀 New app.py generated with correct column mappings.")
    else:
        print("\n❌ Template analysis failed!")
        sys.exit(1)