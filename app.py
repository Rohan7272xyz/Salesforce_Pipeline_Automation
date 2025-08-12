import pandas as pd
import os
import eel
import tkinter as tk
from openpyxl.styles import Alignment
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

# --- EEL Setup ---
eel.init(str(Config.WEB_DIR))

# --- Logging Shim ---
def safe_log(message, level='info'):
    """Logs messages to the eel frontend if available, otherwise prints to console."""
    if hasattr(eel, 'logMessage'):
        try:
            eel.logMessage(message, level)
            return
        except Exception:
            pass
    prefix = level.upper()
    print(f"[{prefix}] {message}")

def safe_complete(ok: bool):
    """Signals completion status to the eel frontend if available."""
    if hasattr(eel, 'processingComplete'):
        try:
            eel.processingComplete(ok)
            return
        except Exception:
            pass
    print(f"[STATUS] processingComplete({ok})")

@eel.expose
def select_file(title, file_type):
    """Opens a native file dialog to select a file."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_path = filedialog.askopenfilename(
            title=title, filetypes=[(file_type, "*.xlsx *.xls")]
        )
        return file_path
    except Exception as e:
        safe_log(f"Error in file selection: {e}", 'error')
        return ""

# --- Date parsing helper ---
def parse_date(val):
    """Safely parses a value to a date object, returning original value on failure."""
    try:
        return pd.to_datetime(val).date()
    except:
        return val

@eel.expose
def process_and_merge_files(data_path):
    """
    Main function to process the raw data file, merge it into a template,
    and save the output. It ensures the 'Total' row is always last.
    Auto-generated with smart column mapping.
    
    Input Data Columns: 11
    Total Template Columns: 45
    
    Smart Column Mapping:
   Capture Manager → Raw Column 1 'Capture Manager'\n   Opportunity Name → Raw Column 3 'Opportunity Name'\n   SF Number → Raw Column 4 'SalesForce ID'\n   T&E → Raw Column 5 'T&E'\n   Stage → Raw Column 6 'Stage'\n   Positioning → Raw Column 7 'Positioning'\n   Ceiling Value ($) → Raw Column 8 'Contract Ceiling Value'\n   MAG Value ($) → Raw Column 9 'MAG Value'\n   Anticipated RFP Date → Raw Column 10 'Anticipated RFP Date'\n   RFP Award → Raw Column 11 'Award Date'\n   GovWin → Raw Column 12 'GovWin IQ Opportunity ID'
    """
    try:
        safe_log("--- Starting Excel Processing ---")

        if not data_path or not Path(data_path).exists():
            safe_log("Error: Raw Data File not found or not provided.", 'error')
            safe_complete(False)
            return None

        if not Config.TEMPLATE_PATH.exists():
            safe_log(f"Error: Integrated template file not found at {Config.TEMPLATE_PATH}", 'error')
            safe_complete(False)
            return None

        # --- Strip Formatting and Metadata ---
        safe_log(f"Reading and cleaning data from '{Path(data_path).name}'...")
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        raw_values = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in sheet_raw.iter_rows(values_only=True)
        ]

        df_raw = pd.DataFrame(raw_values)
        df_raw = df_raw.iloc[14:].copy()  # Data starts at row 14 (0-indexed)
        df_raw.drop(columns=[0], inplace=True, errors='ignore')  # Remove empty first column
        df_raw.reset_index(drop=True, inplace=True)

        # --- Input Data Column Structure (Salesforce fields only) ---
        expected_columns = [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        if any(idx >= len(df_raw.columns) for idx in expected_columns):
            safe_log(f"Error: Expected {len(expected_columns)} columns but input has {len(df_raw.columns)} columns.", 'error')
            safe_log("Input columns needed: " + str(expected_columns), 'error')
            safe_log("Available columns: " + str(list(range(len(df_raw.columns)))), 'error')
            safe_complete(False)
            return None

        df_raw = df_raw[expected_columns]
        df_raw.columns = [
    'Capture Manager',
    'Opportunity Name',
    'SF Number',
    'T&E',
    'Stage',
    'Positioning',
    'Ceiling Value ($)',
    'MAG Value ($)',
    'Anticipated RFP Date',
    'RFP Award',
    'GovWin'
]

        # DEBUG: Check if the problematic columns have data
        safe_log("DEBUG: Column names after mapping:")
        safe_log(str(df_raw.columns.tolist()))
        safe_log("DEBUG: Sample data from first row:")
        for col in df_raw.columns:
            sample_val = df_raw[col].iloc[0] if len(df_raw) > 0 else "NO DATA"
            safe_log(f"  {col}: '{sample_val}'")
        safe_log("DEBUG: Check specific columns:")
        for col_name in ['Award Date', 'GovWin IQ Opportunity ID', 'Positioning']:
            if col_name in df_raw.columns:
                non_empty = df_raw[col_name].dropna()
                non_empty = non_empty[non_empty != '']
                safe_log(f"  {col_name}: {len(non_empty)} non-empty values out of {len(df_raw)}")
                if len(non_empty) > 0:
                    safe_log(f"    Sample values: {non_empty.head(3).tolist()}")
            else:
                safe_log(f"  {col_name}: COLUMN NOT FOUND!")

        # --- Auto-generated Data Processing (Input columns only) ---
        df_raw['Ceiling Value ($)'] = pd.to_numeric(df_raw['Ceiling Value ($)'].str.replace(r'[\$,]', '', regex=True), errors='coerce')
        df_raw['MAG Value ($)'] = pd.to_numeric(df_raw['MAG Value ($)'].str.replace(r'[\$,]', '', regex=True), errors='coerce')

        df = df_raw.dropna(how='all')
        df.dropna(subset=[df.columns[1]], inplace=True)  # Use second column for opportunity check

        # --- Exclude unwanted text across multiple columns ---
        EXCLUSION_KEYWORDS = [
            'Confidential Information - Do Not Distribute',
            'Copyright © 2000-2025 salesforce.com, inc. All rights reserved.'
        ]
        for col in [df.columns[0], df.columns[1]]:  # Check first two columns
            df = df[~df[col].astype(str).str.contains('|'.join(EXCLUSION_KEYWORDS), case=False, na=False)]

        # --- Handle Total Row AFTER filtering ---
        safe_log("Separating and sorting main data from total row...")

        # Create a temporary column for case-insensitive matching and trimming
        df['temp_mgr_lower'] = df[df.columns[0]].astype(str).str.strip().str.lower()

        # Isolate the total row(s) and the main data
        total_row_mask = df['temp_mgr_lower'] == 'total'
        total_df = df[total_row_mask].copy()
        main_df = df[~total_row_mask].copy()

        # Handle stray total count rows
        if not total_df.empty:
            count_row_mask = main_df[df.columns[1]].astype(str).str.match(r'^\d+$')
            if count_row_mask.any():
                numeric_value = main_df.loc[count_row_mask, df.columns[1]].iloc[0]
                total_df.iloc[0, total_df.columns.get_loc(df.columns[1])] = numeric_value
                main_df = main_df[~count_row_mask]
                safe_log("Found and moved stray total count to the total row.")

        # --- Sort data in the desired order: by manager, then unassigned, then total ---
        safe_log("Sorting data: assigned managers, unassigned, then total...")

        has_mgr_mask = (main_df['temp_mgr_lower'] != '') & (main_df['temp_mgr_lower'] != 'nan')
        df_with_mgr = main_df[has_mgr_mask].copy()
        df_without_mgr = main_df[~has_mgr_mask].copy()

        df_with_mgr.sort_values(by=df.columns[0], inplace=True)

        df = pd.concat([df_with_mgr, df_without_mgr, total_df], ignore_index=True)
        df.drop(columns=['temp_mgr_lower'], inplace=True, errors='ignore')

        safe_log(f"Processed {len(df)} rows of data.")

        # DEBUG: Check DataFrame just before writing to Excel
        safe_log("DEBUG: Final DataFrame columns before Excel writing:")
        safe_log(f"  DataFrame shape: {df.shape}")
        safe_log(f"  Columns with data for problem columns:")
        for col_name in ['Award Date', 'GovWin IQ Opportunity ID', 'Positioning']:
            if col_name in df.columns:
                col_index = df.columns.get_loc(col_name)
                non_empty = df[col_name].dropna()
                non_empty = non_empty[non_empty != '']
                safe_log(f"    {col_name} (pandas index {col_index}): {len(non_empty)} values")
            else:
                safe_log(f"    {col_name}: NOT FOUND IN DATAFRAME")

        # --- Load Template ---
        safe_log("Loading integrated template file...")
        workbook = load_workbook(Config.TEMPLATE_PATH)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]

        # --- Clear Old Data (All template columns) ---
        end_row = sheet.max_row
        for r in range(Config.DATA_START_ROW, end_row + 1):
            val = sheet.cell(row=r, column=1).value
            if isinstance(val, str) and "total" in val.lower():
                end_row = r - 1
                break

        # Clear all template columns (input + calendar) - add +1 for correct positioning
        for r_idx in range(Config.DATA_START_ROW, end_row + 1):
            for c_idx in range(2, 47):  # Start from column 2 (B), not 1 (A)
                sheet.cell(row=r_idx, column=c_idx).value = None

        safe_log(f"Cleared rows {Config.DATA_START_ROW} to {end_row} in the template.")

        # --- Write New Data (Only input columns have data) ---
        safe_log("Writing new data to template...")
        for i, row in enumerate(df.itertuples(index=False), start=Config.DATA_START_ROW):
            sheet.row_dimensions[i].height = 30  # Set uniform row height

            for j, val in enumerate(row, start=2):  # START FROM COLUMN 2 (B), NOT 1 (A)
                cell = sheet.cell(row=i, column=j)

                if pd.isna(val):
                    cell.value = None
                elif j == 9:  # Ceiling Value ($)
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif j == 10:  # MAG Value ($)
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif j == 11:  # Anticipated RFP Date
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'
                elif j == 12:  # RFP Award
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'
                else:
                    cell.value = val

                # Text wrapping for specific columns
                if j == 4:  # Opportunity Name
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        downloads_path = Config.get_downloads_path()
        output_filename = Config.DEFAULT_OUTPUT_NAME
        output_path = downloads_path / output_filename

        workbook.save(str(output_path))
        safe_log(f"Success! Final file saved to '{output_path}'.", 'success')
        safe_complete(True)
        return str(output_path)

    except Exception as e:
        safe_log(f"An unexpected error occurred: {e}", 'error')
        import traceback
        safe_log(traceback.format_exc(), 'error')
        safe_complete(False)
        return None

if __name__ == "__main__":
    # Validate configuration first
    try:
        Config.validate_config()
        print("Configuration validated successfully")
    except ValueError as e:
        print(f"Configuration error: {e}")
        sys.exit(1)
    
    # This allows running the script directly for testing or from the command line
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
        result_path = process_and_merge_files(data_path)
        if result_path:
            print(f"[RESULT] {result_path}")
        else:
            sys.exit(1)
    else:
        # This starts the eel GUI application
        try:
            eel.start('index.html', port=0, cmdline_args=['--start-maximized'])
        except (SystemExit, MemoryError, KeyboardInterrupt):
            print("Application closed.")
        except Exception as e:
            print(f"Failed to start GUI: {e}")
            print("Make sure the 'web' directory exists with index.html")
            sys.exit(1)
