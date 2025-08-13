import pandas as pd
import os
import eel
import tkinter as tk
from openpyxl.styles import Alignment
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

# --- EEL Setup ---
eel.init(str(Config.WEB_DIR))

# --- Logging Shim ---
def safe_log(message, level='info'):
    """Logs messages to the eel frontend if available, otherwise prints to console."""
    if hasattr(eel, 'logMessage'):
        try:
            eel.logMessage(message, level)
            return
        except Exception:
            pass
    prefix = level.upper()
    print(f"[{prefix}] {message}")

def safe_complete(ok: bool):
    """Signals completion status to the eel frontend if available."""
    if hasattr(eel, 'processingComplete'):
        try:
            eel.processingComplete(ok)
            return
        except Exception:
            pass
    print(f"[STATUS] processingComplete({ok})")

@eel.expose
def select_file(title, file_type):
    """Opens a native file dialog to select a file."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_path = filedialog.askopenfilename(
            title=title, filetypes=[(file_type, "*.xlsx *.xls")]
        )
        return file_path
    except Exception as e:
        safe_log(f"Error in file selection: {e}", 'error')
        return ""

# --- Dynamic Column Mapper ---
class DynamicColumnMapper:
    """Real-time column mapping that adapts to any input file structure."""
    def __init__(self):
        self.column_matching_rules = {
            'capture manager': {
                'keywords': ['capture', 'manager'],
                'exact_matches': ['capture manager'],
                'data_patterns': [],
                'exclusions': []
            },
            'opportunity name': {
                'keywords': ['opportunity', 'name'],
                'exact_matches': ['opportunity name'],
                'data_patterns': [],
                'exclusions': []
            },
            'sf number': {
                'keywords': ['salesforce', 'sf', 'id'],
                'exact_matches': ['salesforce id', 'sf number'],
                'data_patterns': [],
                'exclusions': []
            },
            'stage': {
                'keywords': ['stage'],
                'exact_matches': ['stage'],
                'data_patterns': ['pre-rfp', 'rfp', 'proposal', 'award'],
                'exclusions': ['equals', 'probability']
            },
            'positioning': {
                'keywords': ['positioning'],
                'exact_matches': ['positioning'],
                'data_patterns': ['sub', 'capture', 'qualification', 'prime'],
                'exclusions': ['equals', 'probability']
            },
            'ceiling value': {
                'keywords': ['ceiling', 'value', 'contract'],
                'exact_matches': ['ceiling value ($)', 'contract ceiling value'],
                'data_patterns': ['$', 'million', 'thousand'],
                'exclusions': ['mag', 'equals', 'probability']
            },
            'mag value': {
                'keywords': ['mag', 'value'],
                'exact_matches': ['mag value ($)', 'mag value'],
                'data_patterns': ['$', 'million', 'thousand'],
                'exclusions': ['contract', 'ceiling', 'equals', 'probability']
            },
            'anticipated rfp date': {
                'keywords': ['anticipated', 'rfp', 'date'],
                'exact_matches': ['anticipated rfp date'],
                'data_patterns': ['/', '-', '2024', '2025'],
                'exclusions': ['award', 'equals', 'probability']
            },
            'rfp award': {
                'keywords': ['award', 'date'],
                'exact_matches': ['rfp award', 'award date'],
                'data_patterns': ['/', '-', '2024', '2025'],
                'exclusions': ['anticipated', 'rfp', 'equals', 'probability']
            },
            'govwin': {
                'keywords': ['govwin', 'iq'],
                'exact_matches': ['govwin iq opportunity id', 'govwin'],
                'data_patterns': [],
                'exclusions': []
            }
        }

    def get_template_structure(self):
        """Get the template column structure."""
        try:
            workbook = load_workbook(Config.TEMPLATE_PATH, data_only=True)
            sheet = workbook[Config.TEMPLATE_SHEET_NAME]

            input_columns = []
            calendar_columns = []

            # Extract template headers starting from column 2
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=4, column=col).value
                if cell_value and str(cell_value).strip():
                    header_name = str(cell_value).strip()
                    header_lower = header_name.lower()

                    # Classify as input or calendar column
                    calendar_keywords = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                         'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                                         'q1', 'q2', 'q3', 'q4', 'quarter', '2024', '2025']

                    is_calendar = any(keyword in header_lower for keyword in calendar_keywords)

                    if is_calendar:
                        calendar_columns.append({'name': header_name, 'index': col})
                    else:
                        input_columns.append({'name': header_name, 'index': col})
                else:
                    if col <= 15:  # Continue for early columns that might be empty
                        continue
                    else:
                        break

            return input_columns, calendar_columns

        except Exception as e:
            safe_log(f"Error getting template structure: {e}", 'error')
            return [], []

    def analyze_raw_data(self, file_path):
        """Analyze the raw data file structure."""
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.sheetnames[0]
            raw_sheet = workbook[sheet]

            # Find header row - look for row with multiple proper column names
            for row_num in range(10, 20):
                try:
                    row_data = list(raw_sheet.iter_rows(values_only=True, min_row=row_num, max_row=row_num))[0]
                    headers = [str(cell).strip() if cell else "" for cell in row_data]

                    # Count how many cells look like proper column headers (not filter text)
                    proper_headers = 0
                    for header in headers:
                        if header and len(header) > 0:
                            # Must be actual column names, not filter text
                            if ('manager' in header.lower() or 
                                'opportunity' in header.lower() or 
                                'salesforce' in header.lower() or
                                'positioning' in header.lower() or
                                'value' in header.lower() or
                                'date' in header.lower() or
                                'govwin' in header.lower()) and                                'equals' not in header.lower() and                                'probability' not in header.lower():
                                proper_headers += 1

                    # Need at least 5 proper column headers to be the header row
                    if proper_headers >= 5:
                        print(f"Found header row at row {row_num} with {proper_headers} proper headers")

                        # Get sample data (next 3 rows)
                        sample_data = []
                        for sample_row in range(row_num + 1, row_num + 4):
                            try:
                                sample_row_data = list(raw_sheet.iter_rows(values_only=True, min_row=sample_row, max_row=sample_row))[0]
                                sample_data.append([str(cell).strip() if cell else "" for cell in sample_row_data])
                            except:
                                break

                        return {
                            'headers': headers,
                            'sample_data': sample_data,
                            'data_start_row': row_num + 1
                        }
                except:
                    continue

            # Fallback to row 14 if no proper header row found
            print("Using fallback row 14 for headers")
            row_data = list(raw_sheet.iter_rows(values_only=True, min_row=14, max_row=14))[0]
            headers = [str(cell).strip() if cell else "" for cell in row_data]
            return {'headers': headers, 'sample_data': [], 'data_start_row': 15}

        except Exception as e:
            safe_log(f"Error analyzing raw data: {e}", 'error')
            return None

    def map_columns(self, template_input_columns, raw_data_info):
        """Map raw data columns to template columns dynamically."""
        if not raw_data_info:
            return {}

        raw_headers = raw_data_info['headers']
        sample_data = raw_data_info.get('sample_data', [])

        mapping = {}

        for template_col in template_input_columns:
            template_name = template_col['name'].lower().strip()
            best_match = None
            best_score = 0

            for raw_idx, raw_header in enumerate(raw_headers):
                if not raw_header or not raw_header.strip():
                    continue

                raw_header_lower = raw_header.lower().strip()
                score = 0

                # Check matching rules
                for rule_key, rules in self.column_matching_rules.items():
                    if any(keyword in template_name for keyword in rule_key.split()):
                        # Exclusions first
                        exclusions = set(rules.get('exclusions', []))
                        if any(exclusion in raw_header_lower for exclusion in exclusions):
                            continue

                        # Exact matches get highest priority
                        if raw_header_lower in rules.get('exact_matches', []) or any(exact in raw_header_lower for exact in rules.get('exact_matches', [])):
                            score += 100

                        # Keyword matches
                        keyword_matches = sum(1 for keyword in rules.get('keywords', []) if keyword in raw_header_lower)
                        if keyword_matches > 0:
                            score += keyword_matches * 20

                        # Special boosts
                        if 'mag' in template_name and 'mag' in raw_header_lower and 'contract' not in raw_header_lower:
                            score += 50
                        if 'award' in template_name and 'award' in raw_header_lower and 'anticipated' not in raw_header_lower:
                            score += 50

                        # Data patterns
                        data_patterns = rules.get('data_patterns', [])
                        if sample_data and data_patterns:
                            sample_values = [row[raw_idx] if raw_idx < len(row) else '' for row in sample_data[:3]]
                            sample_text = ' '.join(str(val).lower() for val in sample_values if val)
                            pattern_matches = sum(1 for pattern in data_patterns if pattern in sample_text)
                            score += pattern_matches * 5

                if score > best_score:
                    best_score = score
                    best_match = {'raw_index': raw_idx, 'raw_header': raw_header, 'score': score}

            if best_match and best_score >= 10:
                mapping[template_col['name']] = best_match
                safe_log(f"Mapped: {template_col['name']} ← Raw Column {best_match['raw_index']} '{best_match['raw_header']}'")

        safe_log(f"Successfully mapped {len(mapping)}/{len(template_input_columns)} columns")
        return mapping

# --- Date parsing helper ---
def parse_date(val):
    """Safely parses a value to a date object, returning original value on failure."""
    try:
        return pd.to_datetime(val).date()
    except:
        return val

@eel.expose
def process_and_merge_files(data_path):
    """
    DYNAMIC processing function that adapts to any input file structure.
    No hardcoded column lists - maps columns in real-time based on what's available.
    """
    try:
        safe_log("--- Starting DYNAMIC Excel Processing ---")

        if not data_path or not Path(data_path).exists():
            safe_log("Error: Raw Data File not found or not provided.", 'error')
            safe_complete(False)
            return None

        if not Config.TEMPLATE_PATH.exists():
            safe_log(f"Error: Template file not found at {Config.TEMPLATE_PATH}", 'error')
            safe_complete(False)
            return None

        # Initialize dynamic mapper
        mapper = DynamicColumnMapper()

        # Get template structure
        template_input_columns, template_calendar_columns = mapper.get_template_structure()
        safe_log(f"Template structure: {len(template_input_columns)} input cols, {len(template_calendar_columns)} calendar cols")

        # Analyze raw data structure
        raw_data_info = mapper.analyze_raw_data(data_path)
        if not raw_data_info:
            safe_log("Error: Could not analyze raw data structure", 'error')
            safe_complete(False)
            return None

        # Create dynamic column mapping
        column_mapping = mapper.map_columns(template_input_columns, raw_data_info)
        if not column_mapping:
            safe_log("Error: No columns could be mapped", 'error')
            safe_complete(False)
            return None

        # --- Load and process raw data ---
        safe_log(f"Reading and cleaning data from '{Path(data_path).name}'...")
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        raw_values = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in sheet_raw.iter_rows(values_only=True)
        ]

        df_raw = pd.DataFrame(raw_values)
        df_raw = df_raw.iloc[14:].copy()  # Data starts at row 14 (0-indexed)
        df_raw.drop(columns=[0], inplace=True, errors='ignore')  # Remove empty first column
        df_raw.reset_index(drop=True, inplace=True)

        # --- DYNAMIC column selection based on mapping ---
        mapped_columns = []
        column_names = []
        raw_to_template_mapping = {}  # Track which raw column goes to which template position

        # Sort template columns by their template position to maintain order
        sorted_template_cols = sorted(template_input_columns, key=lambda x: x['index'])

        for template_col in sorted_template_cols:
            if template_col['name'] in column_mapping:
                raw_index = column_mapping[template_col['name']]['raw_index']
                if raw_index < len(df_raw.columns):
                    mapped_columns.append(raw_index)
                    column_names.append(template_col['name'])
                    raw_to_template_mapping[template_col['name']] = raw_index

        if not mapped_columns:
            safe_log("Error: No valid column mappings found", 'error')
            safe_complete(False)
            return None

        # Keep track of original raw column indices before reindexing
        original_raw_indices = mapped_columns.copy()

        # Select only the mapped columns - this will reindex to 0,1,2,3...
        df_raw = df_raw[mapped_columns]
        df_raw.columns = column_names

        safe_log(f"DYNAMIC mapping applied: {len(mapped_columns)} columns selected")
        for i, (col_name, orig_raw_idx) in enumerate(zip(column_names, original_raw_indices)):
            safe_log(f"  {col_name} ← Raw Column {orig_raw_idx} (now pandas index {i})")

        # Create mapping from template column name to pandas DataFrame index
        template_to_df_index = {name: i for i, name in enumerate(column_names)}

        # --- Process data (currency, dates, etc.) ---
        for col in df_raw.columns:
            if 'value' in col.lower():
                try:
                    df_raw[col] = pd.to_numeric(
                        df_raw[col].astype(str).str.replace(r'[\$,]', '', regex=True),
                        errors='coerce'
                    )
                except Exception:
                    pass

        df = df_raw.dropna(how='all')
        if len(df.columns) > 1:
            df.dropna(subset=[df.columns[1]], inplace=True)

        # --- Exclude unwanted text ---
        EXCLUSION_KEYWORDS = [
            'Confidential Information - Do Not Distribute',
            'Copyright © 2000-2025 salesforce.com, inc. All rights reserved.'
        ]
        for col in df.columns[:2]:  # Check first two columns
            df = df[~df[col].astype(str).str.contains('|'.join(EXCLUSION_KEYWORDS), case=False, na=False)]

        # --- Handle Total Row ---
        safe_log("Separating and sorting main data from total row...")
        df['temp_mgr_lower'] = df[df.columns[0]].astype(str).str.strip().str.lower()

        total_row_mask = df['temp_mgr_lower'] == 'total'
        total_df = df[total_row_mask].copy()
        main_df = df[~total_row_mask].copy()

        # Sort data: managers first, then unassigned, then total
        has_mgr_mask = (main_df['temp_mgr_lower'] != '') & (main_df['temp_mgr_lower'] != 'nan')
        df_with_mgr = main_df[has_mgr_mask].copy()
        df_without_mgr = main_df[~has_mgr_mask].copy()
        df_with_mgr.sort_values(by=df.columns[0], inplace=True)

        df = pd.concat([df_with_mgr, df_without_mgr, total_df], ignore_index=True)
        df.drop(columns=['temp_mgr_lower'], inplace=True, errors='ignore')

        safe_log(f"Processed {len(df)} rows of data.")

        # --- Load Template and write data ---
        safe_log("Loading template and writing data...")
        workbook = load_workbook(Config.TEMPLATE_PATH)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]

        # Clear old data
        end_row = sheet.max_row
        for r in range(Config.DATA_START_ROW, end_row + 1):
            val = sheet.cell(row=r, column=1).value
            if isinstance(val, str) and "total" in val.lower():
                end_row = r - 1
                break

        # Clear template columns
        total_template_cols = len(template_input_columns) + len(template_calendar_columns)
        for r_idx in range(Config.DATA_START_ROW, end_row + 1):
            for c_idx in range(2, total_template_cols + 2):
                sheet.cell(row=r_idx, column=c_idx).value = None

        # Write new data to template - PRESERVE TEMPLATE COLUMN ORDER
        for i, row in enumerate(df.itertuples(index=False), start=Config.DATA_START_ROW):
            sheet.row_dimensions[i].height = 30

            # Write data in template column order, not DataFrame order
            template_col_index = 2  # Start at column B (index 2)

            for template_col in sorted_template_cols:
                if template_col['name'] in column_mapping and template_col['name'] in template_to_df_index:
                    # Get the data from the correct DataFrame column
                    df_col_index = template_to_df_index[template_col['name']]
                    val = row[df_col_index]

                    cell = sheet.cell(row=i, column=template_col_index)

                    if pd.isna(val):
                        cell.value = None
                    elif 'value' in template_col['name'].lower():
                        try:
                            cell.value = float(str(val).replace(',', '').replace('$',''))
                            cell.number_format = '$#,##0'
                        except (ValueError, TypeError):
                            cell.value = val
                    elif 'date' in template_col['name'].lower():
                        cell.value = parse_date(val)
                        cell.number_format = 'mm/dd/yyyy'
                    else:
                        cell.value = val

                    # Text wrapping for opportunity name
                    if template_col_index == 3:  # Opportunity Name column
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                template_col_index += 1

        # Save file
        downloads_path = Config.get_downloads_path()
        output_filename = Config.DEFAULT_OUTPUT_NAME
        output_path = downloads_path / output_filename

        workbook.save(str(output_path))
        safe_log(f"SUCCESS! Dynamic processing complete. File saved to '{output_path}'.", 'success')
        safe_complete(True)
        return str(output_path)

    except Exception as e:
        safe_log(f"An unexpected error occurred: {e}", 'error')
        import traceback
        safe_log(traceback.format_exc(), 'error')
        safe_complete(False)
        return None

if __name__ == "__main__":
    # Validate configuration first
    try:
        Config.validate_config()
        print("Configuration validated successfully")
    except ValueError as e:
        print(f"Configuration error: {e}")
        sys.exit(1)

    # This allows running the script directly for testing
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
        result_path = process_and_merge_files(data_path)
        if result_path:
            print(f"[RESULT] {result_path}")
        else:
            sys.exit(1)
    else:
        # Start GUI
        try:
            eel.start('index.html', port=0, cmdline_args=['--start-maximized'])
        except (SystemExit, MemoryError, KeyboardInterrupt):
            print("Application closed.")
        except Exception as e:
            print(f"Failed to start GUI: {e}")
            print("Make sure the 'web' directory exists with index.html")
            sys.exit(1)
