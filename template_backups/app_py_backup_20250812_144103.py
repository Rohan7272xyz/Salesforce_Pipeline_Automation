import pandas as pd
import os
import eel
import tkinter as tk
from openpyxl.styles import Alignment
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

# --- EEL Setup ---
eel.init(str(Config.WEB_DIR))

# --- Logging Shim ---
def safe_log(message, level='info'):
    """Logs messages to the eel frontend if available, otherwise prints to console."""
    if hasattr(eel, 'logMessage'):
        try:
            eel.logMessage(message, level)
            return
        except Exception:
            pass
    prefix = level.upper()
    print(f"[{prefix}] {message}")

def safe_complete(ok: bool):
    """Signals completion status to the eel frontend if available."""
    if hasattr(eel, 'processingComplete'):
        try:
            eel.processingComplete(ok)
            return
        except Exception:
            pass
    print(f"[STATUS] processingComplete({ok})")

@eel.expose
def select_file(title, file_type):
    """Opens a native file dialog to select a file."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_path = filedialog.askopenfilename(
            title=title, filetypes=[(file_type, "*.xlsx *.xls")]
        )
        return file_path
    except Exception as e:
        safe_log(f"Error in file selection: {e}", 'error')
        return ""

# --- Date parsing helper ---
def parse_date(val):
    """Safely parses a value to a date object, returning original value on failure."""
    try:
        return pd.to_datetime(val).date()
    except:
        return val

def extract_template_headers():
    """Extract headers from template Excel file at row 4."""
    try:
        workbook = load_workbook(Config.TEMPLATE_PATH, data_only=True)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]
        
        headers = []
        # Start from column 2 since column 1 is empty in the template
        for col in range(2, sheet.max_column + 1):
            cell_value = sheet.cell(row=4, column=col).value
            if cell_value and str(cell_value).strip():
                headers.append({
                    'index': col,
                    'name': str(cell_value).strip()
                })
            else:
                # Stop when we hit the first empty cell (non-data columns like calendar)
                if len(headers) >= 10:  # We expect at least 10 data columns
                    break
        
        safe_log(f"📋 Extracted {len(headers)} template headers")
        return headers
        
    except Exception as e:
        safe_log(f"❌ Error extracting template headers: {e}", 'error')
        return None

def classify_template_columns(headers):
    """Separate input data columns from calendar/Gantt columns."""
    input_columns = []
    calendar_columns = []
    
    calendar_keywords = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                       'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                       'q1', 'q2', 'q3', 'q4', 'quarter', '2024', '2025',
                       'gantt', 'calendar', 'timeline']
    
    for header in headers:
        header_lower = header['name'].lower()
        is_calendar = any(keyword in header_lower for keyword in calendar_keywords)
        
        if is_calendar:
            calendar_columns.append(header)
        else:
            input_columns.append(header)
    
    safe_log(f"📊 Template structure: {len(input_columns)} data columns, {len(calendar_columns)} calendar columns")
    return input_columns, calendar_columns

def analyze_raw_data_structure(data_path):
    """Dynamically analyze the structure of the raw input file."""
    try:
        safe_log(f"🔍 Analyzing raw data structure in {Path(data_path).name}")
        
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        # Get first 20 rows to find headers
        raw_values = []
        for row in sheet_raw.iter_rows(values_only=True, min_row=1, max_row=20):
            raw_values.append([str(cell).strip() if cell is not None else "" for cell in row])

        # Find header row (typically row 14)
        header_row_data = None
        for start_row in range(10, len(raw_values)):
            row = raw_values[start_row]
            row_text = " ".join(str(cell).lower() for cell in row if cell and cell.strip())
            
            # Look for key Salesforce column indicators
            indicators = ['capture', 'opportunity', 'salesforce', 'stage', 'positioning', 'govwin']
            matches = sum(1 for indicator in indicators if indicator in row_text)
            
            if matches >= 3:
                safe_log(f"📋 Found raw data headers at row {start_row + 1}")
                header_row_data = row
                data_start_row = start_row + 1
                break
        
        if not header_row_data:
            safe_log("⚠️ Could not identify header row, using row 14 as fallback")
            header_row_data = raw_values[13] if len(raw_values) > 13 else None
            data_start_row = 14
        
        if not header_row_data:
            raise ValueError("Could not find headers in raw data")
        
        # Extract non-empty headers with their column indices
        raw_headers = []
        for i, header in enumerate(header_row_data):
            if header and header.strip():
                raw_headers.append({
                    'index': i,
                    'name': header.strip()
                })
        
        safe_log(f"📊 Found {len(raw_headers)} raw data columns")
        for header in raw_headers:
            safe_log(f"   Column {header['index']}: {header['name']}")
        
        return {
            'headers': raw_headers,
            'data_start_row': data_start_row
        }
        
    except Exception as e:
        safe_log(f"❌ Error analyzing raw data structure: {e}", 'error')
        raise

def smart_column_mapping(template_columns, raw_headers):
    """Create dynamic mapping between template and raw data columns."""
    try:
        safe_log("🔗 Creating smart column mapping...")
        
        mapping = {}
        unmapped_template = []
        unmapped_raw = set(range(len(raw_headers)))
        
        for template_col in template_columns:
            template_name = template_col['name'].lower().strip()
            best_match = None
            best_score = 0
            
            for raw_header in raw_headers:
                raw_name = raw_header['name'].lower().strip()
                score = 0
                
                # Exact name matching
                if template_name == raw_name:
                    score = 100
                elif template_name in raw_name or raw_name in template_name:
                    score = 50
                
                # Specific keyword matching
                keyword_mappings = {
                    'capture': ['capture', 'manager'],
                    'opportunity': ['opportunity', 'name'],
                    'sf': ['salesforce', 'sf', 'id'],
                    'stage': ['stage'],
                    'positioning': ['positioning'],
                    'ceiling': ['ceiling', 'contract', 'value'],
                    'mag': ['mag', 'value'],
                    'rfp': ['rfp', 'anticipated', 'date'],
                    'award': ['award', 'date'],
                    'govwin': ['govwin', 'iq', 'opportunity', 'id']
                }
                
                for key, keywords in keyword_mappings.items():
                    if key in template_name:
                        if any(kw in raw_name for kw in keywords):
                            score += 30
                
                if score > best_score and raw_header['index'] in unmapped_raw:
                    best_score = score
                    best_match = raw_header
            
            if best_match and best_score >= 30:  # Minimum confidence threshold
                mapping[template_col['name']] = best_match
                unmapped_raw.remove(best_match['index'])
                safe_log(f"   ✅ {template_col['name']} → Raw Column {best_match['index']} '{best_match['name']}' (score: {best_score})")
            else:
                unmapped_template.append(template_col['name'])
                safe_log(f"   ❌ No mapping found for: {template_col['name']}")
        
        safe_log(f"📊 Mapping complete: {len(mapping)}/{len(template_columns)} columns mapped")
        
        if unmapped_template:
            safe_log(f"⚠️ Unmapped template columns will be skipped: {', '.join(unmapped_template)}")
        
        return mapping
        
    except Exception as e:
        safe_log(f"❌ Error in column mapping: {e}", 'error')
        raise

@eel.expose
def process_and_merge_files(data_path):
    """
    DYNAMIC version: Analyzes each input file structure and adapts automatically.
    No hardcoded column mappings - fully self-adapting system.
    """
    try:
        safe_log("--- Starting DYNAMIC Excel Processing ---")

        if not data_path or not Path(data_path).exists():
            safe_log("Error: Raw Data File not found or not provided.", 'error')
            safe_complete(False)
            return None

        if not Config.TEMPLATE_PATH.exists():
            safe_log(f"Error: Template file not found at {Config.TEMPLATE_PATH}", 'error')
            safe_complete(False)
            return None

        # Step 1: Analyze template structure
        template_headers = extract_template_headers()
        if not template_headers:
            safe_log("Error: Could not extract template headers", 'error')
            safe_complete(False)
            return None
        
        input_columns, calendar_columns = classify_template_columns(template_headers)
        if not input_columns:
            safe_log("Error: No input data columns found in template", 'error')
            safe_complete(False)
            return None

        # Step 2: Analyze raw data structure
        raw_structure = analyze_raw_data_structure(data_path)
        
        # Step 3: Create dynamic column mapping
        column_mapping = smart_column_mapping(input_columns, raw_structure['headers'])
        
        if not column_mapping:
            safe_log("Error: No column mappings could be established", 'error')
            safe_complete(False)
            return None

        # Step 4: Load and process raw data
        safe_log(f"📥 Loading raw data from {Path(data_path).name}")
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        raw_values = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in sheet_raw.iter_rows(values_only=True)
        ]

        # Create DataFrame starting from data row
        df_raw = pd.DataFrame(raw_values)
        df_raw = df_raw.iloc[raw_structure['data_start_row']:].copy()
        df_raw.reset_index(drop=True, inplace=True)

        # Step 5: Extract only mapped columns in template order
        mapped_data = {}
        for template_col in input_columns:
            if template_col['name'] in column_mapping:
                raw_col_info = column_mapping[template_col['name']]
                raw_col_index = raw_col_info['index']
                
                if raw_col_index < len(df_raw.columns):
                    mapped_data[template_col['name']] = df_raw.iloc[:, raw_col_index]
                else:
                    safe_log(f"⚠️ Warning: Raw column {raw_col_index} not found, skipping {template_col['name']}")
                    mapped_data[template_col['name']] = pd.Series([None] * len(df_raw))
            else:
                # Column not mapped - fill with empty data
                mapped_data[template_col['name']] = pd.Series([None] * len(df_raw))

        # Create final DataFrame with template column order
        df = pd.DataFrame(mapped_data)
        
        # Step 6: Data processing and cleaning
        safe_log("🧹 Cleaning and processing data...")
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove rows where the main identifier (usually opportunity name) is empty
        if len(df.columns) > 1:
            df = df.dropna(subset=[df.columns[1]])  # Usually opportunity name
        
        # Exclude unwanted text
        EXCLUSION_KEYWORDS = [
            'Confidential Information - Do Not Distribute',
            'Copyright © 2000-2025 salesforce.com, inc. All rights reserved.'
        ]
        
        for col in df.columns[:2]:  # Check first two columns
            if col in df.columns:
                df = df[~df[col].astype(str).str.contains('|'.join(EXCLUSION_KEYWORDS), case=False, na=False)]

        # Process currency columns
        for col in df.columns:
            if 'value' in col.lower() and '$' in col.lower():
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(r'[\$,]', '', regex=True), errors='coerce')

        # Step 7: Handle Total Row and Sorting
        safe_log("📊 Organizing data with total row...")
        
        if len(df) > 0 and len(df.columns) > 0:
            # Create temporary column for sorting
            df['temp_mgr_lower'] = df.iloc[:, 0].astype(str).str.strip().str.lower()
            
            # Separate total row from main data
            total_row_mask = df['temp_mgr_lower'] == 'total'
            total_df = df[total_row_mask].copy()
            main_df = df[~total_row_mask].copy()
            
            # Sort main data
            has_mgr_mask = (main_df['temp_mgr_lower'] != '') & (main_df['temp_mgr_lower'] != 'nan')
            df_with_mgr = main_df[has_mgr_mask].copy()
            df_without_mgr = main_df[~has_mgr_mask].copy()
            
            if len(df_with_mgr) > 0:
                df_with_mgr = df_with_mgr.sort_values(by=df_with_mgr.columns[0])
            
            # Recombine data
            df = pd.concat([df_with_mgr, df_without_mgr, total_df], ignore_index=True)
            df = df.drop(columns=['temp_mgr_lower'], errors='ignore')

        safe_log(f"✅ Processed {len(df)} rows of data")

        # Step 8: Write to template
        safe_log("📝 Writing data to template...")
        workbook = load_workbook(Config.TEMPLATE_PATH)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]

        # Clear old data
        end_row = sheet.max_row
        for r in range(Config.DATA_START_ROW, end_row + 1):
            val = sheet.cell(row=r, column=1).value
            if isinstance(val, str) and "total" in val.lower():
                end_row = r - 1
                break

        # Clear data columns (starting from column 2)
        for r_idx in range(Config.DATA_START_ROW, end_row + 1):
            for c_idx in range(2, len(template_headers) + 2):
                sheet.cell(row=r_idx, column=c_idx).value = None

        safe_log(f"🧹 Cleared template rows {Config.DATA_START_ROW} to {end_row}")

        # Write new data
        for i, row in enumerate(df.itertuples(index=False), start=Config.DATA_START_ROW):
            sheet.row_dimensions[i].height = 30
            
            for j, val in enumerate(row, start=2):  # Start from column 2 (B)
                if j - 2 >= len(df.columns):  # Safety check
                    break
                    
                cell = sheet.cell(row=i, column=j)
                col_name = df.columns[j - 2]

                if pd.isna(val):
                    cell.value = None
                elif 'value' in col_name.lower() and '$' in col_name.lower():
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif 'date' in col_name.lower():
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'
                else:
                    cell.value = val

                # Text wrapping for opportunity name
                if j == 3:  # Opportunity name is usually the second data column
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Step 9: Save output
        downloads_path = Config.get_downloads_path()
        output_filename = Config.DEFAULT_OUTPUT_NAME
        output_path = downloads_path / output_filename

        workbook.save(str(output_path))
        safe_log(f"🎉 Success! File saved to '{output_path}'", 'success')
        
        # Log final mapping summary
        safe_log("📋 Final column mapping used:")
        for template_col in input_columns:
            if template_col['name'] in column_mapping:
                raw_info = column_mapping[template_col['name']]
                safe_log(f"   {template_col['name']} ← Raw Column {raw_info['index']} '{raw_info['name']}'")
        
        safe_complete(True)
        return str(output_path)

    except Exception as e:
        safe_log(f"❌ Processing failed: {e}", 'error')
        import traceback
        safe_log(traceback.format_exc(), 'error')
        safe_complete(False)
        return None

if __name__ == "__main__":
    # Validate configuration first
    try:
        Config.validate_config()
        print("Configuration validated successfully")
    except ValueError as e:
        print(f"Configuration error: {e}")
        sys.exit(1)
    
    # This allows running the script directly for testing
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
        result_path = process_and_merge_files(data_path)
        if result_path:
            print(f"[RESULT] {result_path}")
        else:
            sys.exit(1)
    else:
        # Start the eel GUI application
        try:
            eel.start('index.html', port=0, cmdline_args=['--start-maximized'])
        except (SystemExit, MemoryError, KeyboardInterrupt):
            print("Application closed.")
        except Exception as e:
            print(f"Failed to start GUI: {e}")
            print("Make sure the 'web' directory exists with index.html")
            sys.exit(1)