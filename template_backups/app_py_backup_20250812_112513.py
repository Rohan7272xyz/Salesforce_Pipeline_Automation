import pandas as pd
import os
import eel
import tkinter as tk
from openpyxl.styles import Alignment
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from config import Config

# --- EEL Setup ---
eel.init(str(Config.WEB_DIR))

# --- Logging Shim ---
def safe_log(message, level='info'):
    """Logs messages to the eel frontend if available, otherwise prints to console."""
    if hasattr(eel, 'logMessage'):
        try:
            eel.logMessage(message, level)
            return
        except Exception:
            pass
    prefix = level.upper()
    print(f"[{prefix}] {message}")

def safe_complete(ok: bool):
    """Signals completion status to the eel frontend if available."""
    if hasattr(eel, 'processingComplete'):
        try:
            eel.processingComplete(ok)
            return
        except Exception:
            pass
    print(f"[STATUS] processingComplete({ok})")

def debug_column_data(df, step_name, target_columns=None):
    """Debug helper to track specific columns through processing steps."""
    if target_columns is None:
        target_columns = ['Positioning', 'GovWin IQ Opportunity ID', 'Award Date']
    
    safe_log(f"\n=== 🔍 DEBUG CHECKPOINT: {step_name} ===")
    safe_log(f"DataFrame shape: {df.shape}")
    safe_log(f"DataFrame columns: {list(df.columns)}")
    
    for col_name in target_columns:
        # Try different variations of the column name
        possible_names = [col_name]
        if 'Positioning' in col_name:
            possible_names.extend(['positioning', 'Position', 'position', 'Positioning'])
        elif 'GovWin' in col_name:
            possible_names.extend(['GovWin', 'govwin', 'GovWin IQ', 'GovWin ID', 'Gov Win', 'GovWin IQ Opportunity ID'])
        elif 'Award' in col_name:
            possible_names.extend(['Award Date', 'RFP Award', 'Award', 'award date'])
            
        found_col = None
        for possible_name in possible_names:
            if possible_name in df.columns:
                found_col = possible_name
                break
        
        if found_col:
            col_data = df[found_col]
            non_empty = col_data.dropna()
            non_empty = non_empty[non_empty != '']
            non_empty = non_empty[non_empty != 'nan']
            
            safe_log(f"  ✅ FOUND {found_col}:")
            safe_log(f"    - Total values: {len(col_data)}")
            safe_log(f"    - Non-empty values: {len(non_empty)}")
            safe_log(f"    - Data type: {col_data.dtype}")
            
            if len(non_empty) > 0:
                safe_log(f"    - Sample values: {list(non_empty.head(3))}")
                safe_log(f"    - All unique values: {list(non_empty.unique())}")
            else:
                safe_log(f"    - ❌ ALL VALUES ARE EMPTY!")
                
            # Check for specific problematic values
            null_count = col_data.isnull().sum()
            empty_str_count = (col_data == '').sum()
            nan_str_count = (col_data == 'nan').sum()
            
            safe_log(f"    - Null values: {null_count}")
            safe_log(f"    - Empty strings: {empty_str_count}")
            safe_log(f"    - 'nan' strings: {nan_str_count}")
            
            # Show first 5 actual values (including empty ones)
            first_5_values = col_data.head(5).tolist()
            safe_log(f"    - First 5 raw values: {first_5_values}")
                
        else:
            safe_log(f"  ❌ NOT FOUND: {col_name}")
            safe_log(f"    - Available columns: {list(df.columns)}")

def debug_raw_excel_structure(file_path):
    """Debug the raw Excel file structure to understand column positions."""
    safe_log(f"\n=== 📋 RAW EXCEL ANALYSIS: {Path(file_path).name} ===")
    
    try:
        wb_raw = load_workbook(file_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]
        
        # Get all raw values
        raw_values = []
        for row in sheet_raw.iter_rows(values_only=True):
            raw_values.append([str(cell).strip() if cell is not None else "" for cell in row])
        
        safe_log(f"Total rows in Excel: {len(raw_values)}")
        safe_log(f"Total columns in Excel: {len(raw_values[0]) if raw_values else 0}")
        
        # Show row 14 and 15 (where data typically starts)
        for row_idx in [13, 14]:  # 0-indexed, so 13=row 14, 14=row 15
            if row_idx < len(raw_values):
                row_data = raw_values[row_idx]
                safe_log(f"Raw Excel Row {row_idx + 1}: {row_data[:15]}...")  # Show first 15 columns
        
        # Look for Positioning and GovWin in the raw data
        safe_log(f"\n🔍 Searching for Positioning and GovWin in raw Excel...")
        for row_idx, row in enumerate(raw_values[:20]):  # Check first 20 rows
            for col_idx, cell_val in enumerate(row):
                if cell_val and ('positioning' in cell_val.lower() or 'govwin' in cell_val.lower()):
                    safe_log(f"  Found '{cell_val}' at Row {row_idx + 1}, Column {col_idx + 1}")
        
        return raw_values
        
    except Exception as e:
        safe_log(f"❌ Error analyzing raw Excel: {e}")
        return None

@eel.expose
def select_file(title, file_type):
    """Opens a native file dialog to select a file."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_path = filedialog.askopenfilename(
            title=title, filetypes=[(file_type, "*.xlsx *.xls")]
        )
        return file_path
    except Exception as e:
        safe_log(f"Error in file selection: {e}", 'error')
        return ""

# --- Date parsing helper ---
def parse_date(val):
    """Safely parses a value to a date object, returning original value on failure."""
    try:
        return pd.to_datetime(val).date()
    except:
        return val

@eel.expose
def process_and_merge_files(data_path):
    """
    Main function to process the raw data file, merge it into a template,
    and save the output. It ensures the 'Total' row is always last.
    Auto-generated with smart column mapping.
    
    Input Data Columns: 11
    Total Template Columns: 45
    
    Smart Column Mapping:
   Capture Manager → Capture Manager\n   Opportunity Name → Opportunity Name\n   SF Number → SalesForce ID\n   T&E → T&E\n   Stage → Stage\n   Positioning → Positioning\n   Ceiling Value ($) → Contract Ceiling Value\n   MAG Value ($) → MAG Value\n   Anticipated RFP Date → Anticipated RFP Date\n   RFP Award → Award Date\n   GovWin → GovWin IQ Opportunity ID
    """
    try:
        safe_log("--- Starting Excel Processing ---")

        if not data_path or not Path(data_path).exists():
            safe_log("Error: Raw Data File not found or not provided.", 'error')
            safe_complete(False)
            return None

        if not Config.TEMPLATE_PATH.exists():
            safe_log(f"Error: Integrated template file not found at {Config.TEMPLATE_PATH}", 'error')
            safe_complete(False)
            return None

        # ===== DEBUG: Analyze raw Excel structure =====
        raw_excel_data = debug_raw_excel_structure(data_path)

        # --- Strip Formatting and Metadata ---
        safe_log(f"Reading and cleaning data from '{Path(data_path).name}'...")
        wb_raw = load_workbook(data_path, data_only=True)
        sheet_name = wb_raw.sheetnames[0]
        sheet_raw = wb_raw[sheet_name]

        raw_values = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in sheet_raw.iter_rows(values_only=True)
        ]

        df_raw = pd.DataFrame(raw_values)
        
        # ===== DEBUG: Check raw DataFrame before any processing =====
        safe_log(f"\n=== 📊 RAW DATAFRAME (before any processing) ===")
        safe_log(f"Shape: {df_raw.shape}")
        safe_log(f"Column count: {len(df_raw.columns)}")
        if len(df_raw) > 14:
            safe_log(f"Row 15 data (index 14): {df_raw.iloc[14].tolist()[:15]}...")  # Show first 15 columns
        
        # Look for positioning and govwin data in columns 7 and 12 (your specified positions)
        if df_raw.shape[1] > 12:
            safe_log(f"Raw Column 8 (index 7) sample: {df_raw.iloc[14:19, 7].tolist() if len(df_raw) > 14 else 'NO DATA'}")
            safe_log(f"Raw Column 13 (index 12) sample: {df_raw.iloc[14:19, 12].tolist() if len(df_raw) > 14 else 'NO DATA'}")
        
        df_raw = df_raw.iloc[14:].copy()  # Data starts at row 14 (0-indexed)
        
        # ===== DEBUG: After row slicing =====
        debug_column_data(df_raw, "AFTER ROW SLICING (iloc[14:])")
        
        df_raw.drop(columns=[0], inplace=True, errors='ignore')  # Remove empty first column
        
        # ===== DEBUG: After dropping column 0 =====
        debug_column_data(df_raw, "AFTER DROPPING COLUMN 0")
        
        df_raw.reset_index(drop=True, inplace=True)

        # --- Input Data Column Structure (Salesforce fields only) ---
        expected_columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        
        # ===== DEBUG: Column selection analysis =====
        safe_log(f"\n=== 🎯 COLUMN SELECTION ANALYSIS ===")
        safe_log(f"Expected columns: {expected_columns}")
        safe_log(f"Available columns: {list(range(len(df_raw.columns)))}")
        safe_log(f"DataFrame shape before selection: {df_raw.shape}")
        
        # Show what data is in the columns we're about to select
        for idx in expected_columns:
            if idx < len(df_raw.columns):
                sample_data = df_raw.iloc[:3, idx].tolist() if len(df_raw) >= 3 else df_raw.iloc[:, idx].tolist()
                safe_log(f"  Column {idx} sample: {sample_data}")
            else:
                safe_log(f"  Column {idx}: OUT OF RANGE!")
        
        if any(idx >= len(df_raw.columns) for idx in expected_columns):
            safe_log(f"Error: Expected {len(expected_columns)} columns but input has {len(df_raw.columns)} columns.", 'error')
            safe_log("Input columns needed: " + str(expected_columns), 'error')
            safe_log("Available columns: " + str(list(range(1, len(df_raw.columns) + 1))), 'error')
            safe_complete(False)
            return None

        df_raw = df_raw[expected_columns]
        
        # ===== DEBUG: After column selection =====
        debug_column_data(df_raw, "AFTER COLUMN SELECTION")
        
        df_raw.columns = [
    'Capture Manager',
    'Opportunity Name',
    'SalesForce ID',
    'T&E',
    'Stage',
    'Positioning',
    'Contract Ceiling Value',
    'MAG Value',
    'Anticipated RFP Date',
    'Award Date',
    'GovWin IQ Opportunity ID'
]

        # ===== DEBUG: After column renaming =====
        debug_column_data(df_raw, "AFTER COLUMN RENAMING")

        # --- Auto-generated Data Processing (Input columns only) ---
        safe_log(f"\n=== 🔄 DATA PROCESSING PHASE ===")
        
        # Before processing
        debug_column_data(df_raw, "BEFORE DATA PROCESSING")
        
        df_raw['Contract Ceiling Value'] = pd.to_numeric(df_raw['Contract Ceiling Value'].str.replace(r'[\$,]', '', regex=True), errors='coerce')
        df_raw['MAG Value'] = pd.to_numeric(df_raw['MAG Value'].str.replace(r'[\$,]', '', regex=True), errors='coerce')

        # After currency processing
        debug_column_data(df_raw, "AFTER CURRENCY PROCESSING")

        df = df_raw.dropna(how='all')
        
        # After dropping all-NA rows
        debug_column_data(df, "AFTER DROPPING ALL-EMPTY ROWS")
        
        df.dropna(subset=[df.columns[1]], inplace=True)  # Use second column for opportunity check

        # After dropping rows with empty opportunity names
        debug_column_data(df, "AFTER DROPPING EMPTY OPPORTUNITY ROWS")

        # --- Exclude unwanted text across multiple columns ---
        EXCLUSION_KEYWORDS = [
            'Confidential Information - Do Not Distribute',
            'Copyright © 2000-2025 salesforce.com, inc. All rights reserved.'
        ]
        
        initial_row_count = len(df)
        for col in [df.columns[0], df.columns[1]]:  # Check first two columns
            df = df[~df[col].astype(str).str.contains('|'.join(EXCLUSION_KEYWORDS), case=False, na=False)]
        
        if len(df) != initial_row_count:
            safe_log(f"Excluded {initial_row_count - len(df)} rows with unwanted text")
        
        # After exclusion filtering
        debug_column_data(df, "AFTER EXCLUSION FILTERING")

        # --- Handle Total Row AFTER filtering ---
        safe_log("Separating and sorting main data from total row...")

        # Create a temporary column for case-insensitive matching and trimming
        df['temp_mgr_lower'] = df[df.columns[0]].astype(str).str.strip().str.lower()

        # Isolate the total row(s) and the main data
        total_row_mask = df['temp_mgr_lower'] == 'total'
        total_df = df[total_row_mask].copy()
        main_df = df[~total_row_mask].copy()

        # Debug total row separation
        safe_log(f"Total rows found: {len(total_df)}")
        safe_log(f"Main data rows: {len(main_df)}")
        
        debug_column_data(main_df, "MAIN DATA (after total separation)")
        if len(total_df) > 0:
            debug_column_data(total_df, "TOTAL ROW DATA")

        # Handle stray total count rows
        if not total_df.empty:
            count_row_mask = main_df[df.columns[1]].astype(str).str.match(r'^\d+$')
            if count_row_mask.any():
                numeric_value = main_df.loc[count_row_mask, df.columns[1]].iloc[0]
                total_df.iloc[0, total_df.columns.get_loc(df.columns[1])] = numeric_value
                main_df = main_df[~count_row_mask]
                safe_log("Found and moved stray total count to the total row.")

        # --- Sort data in the desired order: by manager, then unassigned, then total ---
        safe_log("Sorting data: assigned managers, unassigned, then total...")

        has_mgr_mask = (main_df['temp_mgr_lower'] != '') & (main_df['temp_mgr_lower'] != 'nan')
        df_with_mgr = main_df[has_mgr_mask].copy()
        df_without_mgr = main_df[~has_mgr_mask].copy()

        df_with_mgr.sort_values(by=df.columns[0], inplace=True)

        df = pd.concat([df_with_mgr, df_without_mgr, total_df], ignore_index=True)
        df.drop(columns=['temp_mgr_lower'], inplace=True, errors='ignore')

        safe_log(f"Processed {len(df)} rows of data.")

        # ===== FINAL DEBUG: Just before Excel writing =====
        debug_column_data(df, "FINAL DATA (ready for Excel writing)")

        # --- Load Template ---
        safe_log("Loading integrated template file...")
        workbook = load_workbook(Config.TEMPLATE_PATH)
        sheet = workbook[Config.TEMPLATE_SHEET_NAME]

        # --- Clear Old Data (All template columns) ---
        end_row = sheet.max_row
        for r in range(Config.DATA_START_ROW, end_row + 1):
            val = sheet.cell(row=r, column=1).value
            if isinstance(val, str) and "total" in val.lower():
                end_row = r - 1
                break

        # Clear all template columns (input + calendar)
        for r_idx in range(Config.DATA_START_ROW, end_row + 1):
            for c_idx in range(1, 46):
                sheet.cell(row=r_idx, column=c_idx).value = None

        safe_log(f"Cleared rows {Config.DATA_START_ROW} to {end_row} in the template.")

        # --- Write New Data (Only input columns have data) ---
        safe_log("Writing new data to template...")
        safe_log(f"\n=== 📝 EXCEL WRITING PHASE ===")
        
        for i, row in enumerate(df.itertuples(index=False), start=Config.DATA_START_ROW):
            sheet.row_dimensions[i].height = 30  # Set uniform row height

            for j, val in enumerate(row, start=1):
                cell = sheet.cell(row=i, column=j)

                # Debug specific columns being written
                if j in [6, 11]:  # Positioning (6) and GovWin (11) based on the column mapping
                    safe_log(f"Writing to Excel row {i}, col {j}: '{val}' (type: {type(val)})")

                if pd.isna(val):
                    cell.value = None
                elif j == 8:  # Ceiling Value ($)
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif j == 9:  # MAG Value ($)
                    try:
                        cell.value = float(val)
                        cell.number_format = '$#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif j == 10:  # Anticipated RFP Date
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'
                elif j == 11:  # RFP Award
                    cell.value = parse_date(val)
                    cell.number_format = 'mm/dd/yyyy'
                else:
                    cell.value = val

                # Text wrapping for specific columns
                if j == 3:  # Opportunity Name
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        downloads_path = Config.get_downloads_path()
        output_filename = Config.DEFAULT_OUTPUT_NAME
        output_path = downloads_path / output_filename

        workbook.save(str(output_path))
        safe_log(f"Success! Final file saved to '{output_path}'.", 'success')
        safe_complete(True)
        return str(output_path)

    except Exception as e:
        safe_log(f"An unexpected error occurred: {e}", 'error')
        import traceback
        safe_log(traceback.format_exc(), 'error')
        safe_complete(False)
        return None

if __name__ == "__main__":
    # Validate configuration first
    try:
        Config.validate_config()
        print("Configuration validated successfully")
    except ValueError as e:
        print(f"Configuration error: {e}")
        sys.exit(1)
    
    # This allows running the script directly for testing or from the command line
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
        result_path = process_and_merge_files(data_path)
        if result_path:
            print(f"[RESULT] {result_path}")
        else:
            sys.exit(1)
    else:
        # This starts the eel GUI application
        try:
            eel.start('index.html', port=0, cmdline_args=['--start-maximized'])
        except (SystemExit, MemoryError, KeyboardInterrupt):
            print("Application closed.")
        except Exception as e:
            print(f"Failed to start GUI: {e}")
            print("Make sure the 'web' directory exists with index.html")
            sys.exit(1)